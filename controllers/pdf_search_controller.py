import os
from typing import Any

import pandas as pd
from PySide6.QtWidgets import QMessageBox, QFileDialog
from openpyxl.styles import PatternFill, Font, Alignment
from views.table_builder import TableBuilder
from services.notification_service import NotificationServices

class PDFSearchController:
    def __init__(self, ui):
        self.ui = ui
        self.worker = None

        self.file_controller = None

        self.matched_results = []
        self.highlighted_results = []

        self.desktop_notification = NotificationServices()
        self.search_complete = False
        self.initial_stat()

    def start_search(self) -> None:
        """ clear previous search results and perform a new one.  """
        from workers.pdf_search_worker import PDFSearchWorker

        self.clear_pdfs_results()

        if not self.file_controller.selected_pdfs:
            self.file_controller.select_pdf_files_path()

        selected_pdfs = self.file_controller.selected_pdfs
        mode_search = self._get_search_mode()
        search_text = self.ui.Fetch_pdf_led.text().strip()

        if not selected_pdfs:
            self.ui.pdf_ptext.appendPlainText("No PDF files selected for search.")
            return

        if mode_search == "Matched Text":
            if not search_text:
                self.ui.pdf_ptext.appendPlainText("Please enter a search term before searching.")
                return
            self.worker = PDFSearchWorker(file_paths=selected_pdfs, mode_search=mode_search, search_text=search_text)

        elif mode_search == "Highlighted Text":
            self.worker = PDFSearchWorker(file_paths=selected_pdfs, mode_search=mode_search)

        else:
            self.ui.pdf_ptext.appendPlainText("Invalid search mode selected.")
            return

        self.search_complete = False

        search_target = "PDFs only" if mode_search == "Highlighted text" else "PDFs and input text"
        self.ui.pdf_ptext.appendPlainText(f"Searching for '{search_text}' in {len(selected_pdfs)} {search_target}...")

        # Connect worker signals
        self.worker.progress.connect(self._update_pdf_search_progress)
        self.worker.result.connect(lambda fp, fn, pn, content, rtype:
                                   TableBuilder.add_result_to_tree(self.ui.pdfs_results_tree, self.matched_results,
                                                                   self.highlighted_results,
                                                                   fp, fn, pn, content, rtype))
        self.worker.finished.connect(self._handle_search_finished)
        self.worker.error_occurred.connect(lambda err: self._append_pdf_log(err))
        self.worker.start()

    def _handle_search_finished(self) -> None:
        """Handle the search end"""
        self.search_complete = True
        self._append_pdf_log("PDF Search Completed.")

    def _handle_error(self, message: str) -> None:
        """Display Error messages"""
        QMessageBox.critical(self.ui.centralwidget, "Error", message)

    def export_results(self) -> None:
        """ Exports all found search results to an Excel file. """
        self.ui.pdf_ptext.clear()
        if not self.search_complete:
            QMessageBox.warning(self.ui.centralwidget, "Export Error", "Complete a search first!")
            return

        try:
            if not self.matched_results and not self.highlighted_results:
                self._handle_error("No results to export")
                return

            current_mode = self._get_search_mode()
            data = self.matched_results if current_mode == "Matched Text" else self.highlighted_results
            sheet_name = current_mode

            self._append_pdf_log(f"Exporting {len(data)} results for mode: {sheet_name}")

            sanitized_data = [
                [
                    self._sanitize_excel_text(item) if isinstance(item, str) else item
                    for item in row
                ]
                for row in data
            ]

            df = pd.DataFrame(sanitized_data, columns=["File Name", "Page", current_mode, "File Path"])
            df.sort_values(by=['File Name', 'Page'], inplace=True)

            path, _ = QFileDialog.getSaveFileName(self.ui.centralwidget, "Save Results", "", "Excel Files (*.xlsx)")
            if not path:
                return

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                header_fill = PatternFill(start_color="3f3f3f", fill_type="solid")
                header_font = Font(bold=True, color="000000")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                worksheet.column_dimensions['A'].width = 30  # File Name
                worksheet.column_dimensions['B'].width = 10  # Page
                worksheet.column_dimensions['C'].width = 80  # Highlighted/Matched Text
                worksheet.column_dimensions['D'].width = 50  # File Path

                wrap_alignment = Alignment(wrap_text=True, vertical='top')
                for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
                    for cell in row:
                        cell.alignment = wrap_alignment

                worksheet.freeze_panes = "A2"

                for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=4):
                    cell = row[0]
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

            QMessageBox.information(self.ui.centralwidget, "Export Successful", f"Results exported to: {path}")
            os.startfile(path)

        except Exception as e:
            self._handle_error(f"Export failed: {str(e)}")

    @staticmethod
    def _sanitize_excel_text(text) -> str | Any:
        """ Cleans extracted text for safe Excel export """
        import re

        if not isinstance(text, str):
            return text  # Only process strings

        # Remove invalid control characters (except common line breaks)
        text = re.sub(r'[\x00-\x08\x0B-\x1F\x7F]', '', text)

        # Remove excessive spaces and formatting issues
        text = re.sub(r'\s+', ' ', text).strip()

        # Ensure text does not start with "=" to prevent formula injection
        if text.startswith(('=', '+', '-', '@')):
            text = "'" + text  # Prefix with an apostrophe to make it plain text in Excel

        return text

    def _update_pdf_search_progress(self, message, count, percent)-> None:
        """the summary of the search result """
        self._append_pdf_log(f"{count} Matches - {message} [{percent}%]")

    def _append_pdf_log(self, message)-> None:
        """Display messages"""
        self.ui.pdf_ptext.appendPlainText(message)

    def initial_stat(self) -> None:
        """the initial status of the pdf section"""
        line_search = self.ui.Fetch_pdf_led
        current_mode = self.ui.fetch_pdf_mode_cbox.currentText()
        if current_mode == "Highlighted Text":
            line_search.setReadOnly(True)
        else:
            line_search.setReadOnly(False)
        self.clear_pdfs_results()

    def clear_pdfs_results(self) -> None:
        """Clear the content and data lists"""
        self.ui.pdfs_results_tree.clear()
        self.ui.pdf_ptext.clear()
        self.matched_results.clear()
        self.highlighted_results.clear()
        self.ui.Fetch_pdf_led.clear()

    def _get_search_mode(self) -> None|str:
        mode_translations = {
            "Texte surligné": "Highlighted Text",
            "النص المميز": "Highlighted Text",
            "النص المطابق": "Matched Text",
            "Texte correspondant": "Matched Text",
            "Matched Text": "Matched Text",
            "Highlighted Text": "Highlighted Text"

        }

        selected_mode = self.ui.fetch_pdf_mode_cbox.currentText().strip()
        return mode_translations.get(selected_mode, None)